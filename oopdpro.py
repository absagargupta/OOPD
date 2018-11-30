'''
Arunav Dutta MT18156
Bandana Prasad MT18158
Sagar Gupta MT18174
'''

from openpyxl import load_workbook
import pandas as pd


Doctor = r'Doctors_1.XLSX'
Patient = r'Patients_1.xlsx'
Doctor1 = r'Doctors_1.xlsx'
Patients1 = r'Patients_1.xlsx'
Var1 = pd.read_excel(Doctor1)
Var2 = pd.read_excel(Patients1)


class Doctor_registration:
    def __init__(self,name,email_id,phone_no,address,schedule,age,password,position,specialization,department,surgeon,timings2):
        self.name=name
        self.email_id=email_id
        self.phone_no=phone_no
        self.address=address
        self.schedule=schedule
        self.age=age
        self.password=password
        self.position=position
        self.department=department
        self.specialization=specialization
        self.surgeon=surgeon
        self.timing2s = timings2
    def save_doctor(self):
        print(self.name)
        wb2 = load_workbook('Doctors_1.xlsx')
        cell1=wb2.active
        df = pd.DataFrame(cell1.values,columns=['name','email_id','phone_no','address','schedule','age','password','position','specialization','department','surgeon','timings2'])
        data1={'name':[self.name],'email_id':[self.email_id],'phone_no':[self.phone_no],'address':[self.address],'schedule':[self.schedule],'age':[self.age],'password':[self.password],'position':[self.position],'specialization':[self.specialization],'department':[self.department],'surgeon':[self.surgeon],'timings2':[self.timings2]}
        df1=pd.DataFrame(data1,columns=['name','email_id','phone_no','address','schedule','age','password','position','specialization','department','surgeon','timings2'])
        df2=pd.concat([df,df1])
        writer = pd.ExcelWriter('Doctors_1.xlsx')
        df2.to_excel(writer,'Sheet1',index=False,index_label=False,header=False)
        writer.save()
'''
class user_registration:
    def __init__(self,user_name,Age,Contact):
        self.user_name=user_name
        self.Age=Age
        self.Contact=Contact
    
    def save_user(self):
        print(self.user_name)
        wb2 = load_workbook('MMT.xlsx')
        cell1=wb2.active
        df = pd.DataFrame(cell1.values,columns=['Name','Age','Contact'])
        data1={'Name':[self.user_name],'Age':[self.Age],'Contact':[self.Contact]}
        df1=pd.DataFrame(data1,columns=['Name','Age','Contact'])
        df2=pd.concat([df,df1])
        writer = pd.ExcelWriter('MMT.xlsx')
        df2.to_excel(writer,'Sheet1',index=False,index_label=False,header=False)
        writer.save()
'''        
'''  
 elif (companyoruser==2):
            name=input("Enter the name of the Doctor")
            email_id=input("Input email ID")
            phone_no=input("Input the contact number")
            address = input("Input the address")
            schedule=input("input the schedule you wish to choose")
            age=input("input the age")
            password=input("input the password")
            position=input("input the position")
            specialization=input("input the specialization")
            department=input("input the department you wish to apply to")
            new=Doctor_registration(name,email_id,phone_no,address,schedule,age,password,position,specialization,department)
            new.save_doctor

      
class doctor_registration:
    def __init__(self,company_name,company_contact,mode_offered,cost):
        self.company_name=company_name
        self.company_contact=company_contact
        self.mode_offered=mode_offered
        self.cost=cost
    
    def save_company(self):
        print(self.company_name)
        wb2 = load_workbook('Company.xlsx')
        cell1=wb2.active
        df = pd.DataFrame(cell1.values,columns=['company_name','company_contact','mode_offered','cost'])
        data1={'company_name':[self.company_name],'company_contact':[self.company_contact],'mode_offered':[self.mode_offered],'cost':[self.cost]}
        df1=pd.DataFrame(data1,columns=['company_name','company_contact','mode_offered','cost'])
        df2=pd.concat([df,df1])
        writer = pd.ExcelWriter('Company.xlsx')
        df2.to_excel(writer,'Sheet1',index=False,index_label=False,header=False)
        writer.save()
        
       
   '''     
class Patient_Registration:
    def __init__(self,name,patient_ID,age,email,password,phone,requirement,location,assigned_doctor,visiting_date,discharging_date,prescribed_medicine,test,disease_identified):
        self.name=name
        self.patient_ID = patient_ID
        self.age=age
        self.email=email
        self.password=password
        self.phone=phone
        self.requirement=requirement
        self.location=location
        self.assigned_doctor=assigned_doctor
        self.visiting_date=visiting_date
        self.discharging_date=discharging_date
        self.prescribed_medicine=prescribed_medicine
        self.test=test
        self.disease_identified=disease_identified
        

    
    def save_patient(self):
        print(self.name)
        wb2 = load_workbook('Patients_1.xlsx')
        cell1=wb2.active
        df = pd.DataFrame(cell1.values,columns=['name','Patient_ID','age','email-ID','password','number','requirement','assigned_doctor','visiting_date','discharge_date','prescribed_medicines','recommended_test','disease_identified'])
        data1={'name':[self.name],'email_id':[self.Patient_ID],'age':[self.age],'email-ID':[self.email],'password':[self.password],'number':[self.number],'requirement':[self.requirement],'assigned_doctor':[self.assigned_doctor],'visiting_date':[self.visiting_date],'discharge_date':[self.discharge_date],'prescribed_medicines':[self.prescribed_medicines],'recommended_test':[self.recommended_test],'disease_identified':[self.disease_identified]}
        df1=pd.DataFrame(data1,columns=['name','Patient_ID','age','email-ID','password','number','requirement','assigned_doctor','visiting_date','discharge_date','prescribed_medicines','recommended_test','disease_identified'])
        df2=pd.concat([df,df1])
        writer = pd.ExcelWriter('Patients.xlsx')
        df2.to_excel(writer,'Sheet1',index=False,index_label=False,header=False)
        writer.save()
        
    def patient_reg(self):### inputted by the patient
        wb2 = load_workbook('Patients_1.xlsx')
        cell1=wb2.active
        print(self.name)
        wb3 = load_workbook('Patients_1.xlsx')
        cell1=wb3.active
        df = pd.DataFrame(cell1.values,columns=['name','-','age','-','-','-','requirement','-','-','-','-','-','-'])
        data1={'name':[self.name],'age':[self.age],'email-ID':[self.email],'requirement':[self.requirement]}
        df1=pd.DataFrame(data1,columns=['name','age','requirement'])
        df2=pd.concat([df,df1])
        writer = pd.ExcelWriter('Patients_1.xlsx')
        df2.to_excel(writer,'Sheet1',index=False,index_label=False,header=False)
        writer.save()
    
    
    def check_patient(self):
        print(self.name)
        wb2 = load_workbook('Patients_1.xlsx')
        cell1=wb2.active
        df = pd.DataFrame(cell1.values,columns=['name','email_id','phone_no','address','schedule','age','password','position','specialization','department'])
        data1={'name':[self.name],'email_id':[self.email_id],'phone_no':[self.phone_no],'address':[self.address],'schedule':[self.schedule],'age':[self.age],'password':[self.password],'position':[self.position],'specialization':[self.specialization],'department':[self.department]}
        df1=pd.DataFrame(data1,columns=['name','email_id','phone_no','schedule','age','password','position','specialization','department'])
        
        
        
class Register:
    def __init__(self,name,age,phone,email_id,password):
        self.name=name
        self.age=age
        self.phone=phone
        self.email_id=email_id
        self.password=password


class patient_Login:
    def __init__(self,name,Id):
        self.name=name
        self.Id=Id
    def patient_chk(self):
        wb2 = load_workbook('Patients_1.xlsx')
        cell1=wb2.active
        df = pd.DataFrame(cell1.values,columns=['Name','ID','age','email-ID','number','ailment','location','doctor_assigned','visiting_date','discharge_date','prescribed_medicines','recommended_test','diagnosis','department'])
        print (df.loc[:,'Name'])
        #print (df)
        
class doctor_Login:
    def __init__(self,dr_login_name,dr_email_ID):
        self.dr_login_name=dr_login_name
        self.dr_email_ID=dr_email_ID
    def dr_chk(self):
        wb2 = load_workbook('Doctors_1.xlsx')
        cell1=wb2.active
        df = pd.DataFrame(cell1.values,columns=['dr_login_name','dr_email_ID','phone_no','address','schedule','age','password','position','specialization','department','surgeon','timings2'])
        print (df.loc[:,'dr_login_name'])
        #print (df)

class Admin:
    def __init__(self,name,email_id,password):
        self.name=name
        self.email_id=email_id
        self.password=password
class Support_staff:
    def __init__(self,name,age,phone_no,position):
        self.name=name
        self.age=age
        self.phone_no=phone_no
        self.position=position
class Hospital:
    def __init__(self,Department_list,doctors_list,patients_list):
        self.Department_list=Department_list
        self.doctors_list=doctors_list
        self.patients_list=patients_list
class Department:    
    def __init__(self,name,id,no_of_doctors,HOD,list_of_doctors):
        self.name=name
        self.no_of_doctors=no_of_doctors
        self.HOD=HOD
        self.list_of_doctors=list_of_doctors

def main():
    
    choice = int(input('Enter Choice:\n1 patient Menu\n2. Doctor Menu'))
    
    
            
    if(choice==1):       
        choice3 = int(input('Enter Choice:\n1 Login\n2. Register'))
        if(choice3==1):
            patient_name=input("name: ")
            patient_ID=input("ID")
            p_emailID = input("Email-ID:")
            p_contact_number = input("Enter your contact number")
            user=patient_Login(patient_name,patient_ID)
            user.patient_chk()
            print("you have been logged in successfully")
            
            
        if(choice3==2):
            p_name = input("name: ")
            p_age = input("age: ")
            p_emailID = input("Email-ID:")
            p_contact_number = input("Enter your contact number")
       
        
        
        print("If you know the department you wish to be treated in then press 1 otherwise 2")
        choice32 = int(input('Enter Choice:\n1 I know the department\n2. I dont know the department'))
        if (choice32 == 1):
           choice42 = input('Enter choice:\n1 Neurology\n2 Cardiology\n3 ENT\n4 Orthopaedics\n5 Physician')
           chosen_department = choice42
           df11 = Var1
           print(df11.loc[df11['Department'] == str(chosen_department)])
           
           qwe = input("Enter the serial number of the doctor you wish to choose")
           chosen_doctor = Var1.iloc[int(qwe)][0]
           print("You have chosen Dr", chosen_doctor, "of", chosen_department)
        if (choice32 == 2):
            choice43 = int(input("Enter one of the ailments of the patient: \n1 Blocked Vessel Constriction \n2Stressed \n3 Vessel Burst\n4 Stroke \n5 Heart attack\n6 Palpitation \n7 Q-tip in ear\n8 Nose Bleeding\n9 Cough or cold\n10 Pain in ear or nose or throat\n11 Broken Bone\n12 knee or hip replacement\n13 broken leg or arm\n14 High Blood Pressure\n15 High Pulse\n16 Food Poisoning or diarheaa ")) 
            if (choice43 == 1 or 2 or 3 or 4):
                chosen_department = 'Neurology'
            if (choice43 == 5 or 6):
                chosen_department = 'Cardiology'
            if (choice43 == 7 or 8 or 9 or 10):
                chosen_department = 'ENT'
            if (choice43 == 11 or 12 or 13):
                chosen_department = 'Orthopaedic'
            if (choice43 == 14 or 15 or 16):
                chosen_department = 'Physician'
            print("You have been alloted ", chosen_department)
            print("choose a doctor from your chosen department")
            print(df11.loc[df11['Department'] == str(chosen_department)])
            qwe1 = input("Enter the serial number of the doctor you wish to choose")
            chosen_doctor1 = Var1.iloc[int(qwe1)][0]
            print("You have chosen Dr", chosen_doctor1, "of", chosen_department)
        
        ###### slot booking
        k1 = 0
        k2 = 0
        k3 = 0
        k4 = 0
        k5 = 0
        k6 = 0
        k7 = 0
        k8 = 0
        k9 = 0
        k10 = 0
        k11 = 0
        k12=0
        choice2213 = int(input("Enter Choice:\n1.Slot_booking \n2.Patient Registration number allotment"))
        if (choice2213 == 1):
                print("The overall timing slots are : ")
                for i in range(0,3):
                    print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+((i+1)*15))
                print(Var1.iloc[0][4]+45,'-',Var1.iloc[0][4]+59)    
                for k in range(0,3):
                    print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+((k+1)*15))
                print(Var1.iloc[0][4]+145,'-',Var1.iloc[0][4]+159) 
                choi=int(input("Enter your slot of choice : "))
                
                if(choi==1):
                    if(k9==0):
                        print("Your slot will be 8am-8:15am")
                        k9=k9+1
                    elif(k9==1):
                        print("Please choose another slot")
                        print("The available slots are : ")
                        for i in range(1,3):
                            print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+((i+1)*15))   
                        print(Var1.iloc[0][4]+45,'-',Var1.iloc[0][4]+59) 
                        for k in range(0,3):
                            print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+((k+1)*15))
                        print(Var1.iloc[0][4]+145,'-',Var1.iloc[0][4]+159)
                        
                elif(choi==2):
                    if(k1==0):
                        print("Your slot will be 8:15am-8:30am")
                        k1=k1+1
                    elif(k1==1):
                        print("Please choose another slot")  
                        print("The available slots are : ")
                        for i in range(0,3,2):
                        
                            print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+(i+1)*15)
                        
                        print(Var1.iloc[0][4]+45,'-',Var1.iloc[0][4]+59) 
                        for k in range(0,3):
                            print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+((k+1)*15))
                        print(Var1.iloc[0][4]+145,'-',Var1.iloc[0][4]+159)     
                        
                elif(choi==3):
                    if(k2==0):
                        print("Your slot will be 8:30am-8:45am") 
                        k2=k2+1
                    elif(k2==1):
                        print("Please choose another slot")
                        for i in range(0,2):
                            print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+(i+1)*15)   
                        print(Var1.iloc[0][4]+45,'-',Var1.iloc[0][4]+59) 
                        for k in range(0,3):
                            print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+(k+1)*15)
                        print(Var1.iloc[0][4]+145,'-',Var1.iloc[0][4]+159)      
                              
                elif(choi==4):
                    if(k3==0):
                        print("Your slot will be 8:45am-9am")
                        k3=k3+1
                    elif(k3==1):
                        print("Please choose another slot")
                        for i in range(0,3):
                            print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+(i+1)*15)   
                        for k in range(0,3):
                            print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+(k+1)*15)
                        print(Var1.iloc[0][4]+145,'-',Var1.iloc[0][4]+159)   
                        
                elif(choi==5):
                    if(k4==0):
                        print("Your slot will be 9am-9:15am")
                        k4=k4+1
                    elif(k4==1):
                        print("Please choose another slot")
                        for i in range(0,3):
                            print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+(i+1)*15)  
                        print(Var1.iloc[0][4]+45,'-',Var1.iloc[0][4]+59) 
                        for k in range(1,3):
                            print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+(k+1)*15)
                        print(Var1.iloc[0][4]+145,'-',Var1.iloc[0][4]+159)
                              
                elif(choi==6):
                    if(k5==0):
                        print("Your slot will be 9:15am-9:30am")
                        k5=k5+1
                    elif(k5==1):
                        print("Please choose another slot")   
                        for i in range(0,3):
                            print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+(i+1)*15)  
                        print(Var1.iloc[0][4]+45,'-',Var1.iloc[0][4]+59) 
                        for k in range(0,3,2):
                            print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+(k+1)*15)
                        print(Var1.iloc[0][4]+145,'-',Var1.iloc[0][4]+159)
                              
                elif(choi==7):
                    if(k6==0):
                        print("Your slot will be 9:30am-9:45am")
                        k6=k6+1
                    elif(k6==1):
                        print("Please choose another slot")  
                        for i in range(0,3):
                            print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+(i+1)*15)   
                        print(Var1.iloc[0][4]+45,'-',Var1.iloc[0][4]+59) 
                        for k in range(0,2):
                            print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+(k+1)*15)
                        print(Var1.iloc[0][4]+145,'-',Var1.iloc[0][4]+159)
                              
                elif(choi==8):
                    if(k7==0):
                        print("Your slot will be 9:45am-10am")
                        k7=k7+1
                    elif(k7==1):
                        print("Please choose another slot")    
                        for i in range(0,3):
                            print(Var1.iloc[0][4]+(i*15) , '-', Var1.iloc[0][4]+(i+1)*15)   
                        print(Var1.iloc[0][4]+45,'-',Var1.iloc[0][4]+59) 
                        for k in range(0,3):
                            print(Var1.iloc[0][4]+100+(k*15),'-', Var1.iloc[0][4]+100+(k+1)*15)
        
        if(choice==2):
            Var2 = pd.read_excel(Patients1)
            print(Var2)
            print(Var2.iloc[0][13])
            #above lines for testing purposes
            
            #roll no. allocation
            if(chosen_department == "Neurology"):
                k8=k8+1
                reg_number = ("NEUR_"+str(k8))
                reg_number = print("Your roll no. is : NEUR_",k8)
            elif(chosen_department =="Cardiology"):
                k9=k9+1
                reg_number = ("NEUR_"+str(k9))
                print("Your roll no. is : CARD_",k9)
            elif(chosen_department == "ENT"):
                k10=k10+1
                reg_number = ("CARD_"+str(k10))
                print("Your roll no. is : ENT_",k10)
            elif(chosen_department == "Orthopaedics"):
                k11=k11+1
                reg_number = ("ORTHO_"+str(k11))
                print("Your roll no. is : ORTHO_",k11)
            elif(chosen_department == "Physician"):
                k12=k12+1
                reg_number = ("PHY_"+str(k12))
                print(str("Your roll no. is : PHY_"+str(k12)))
                
        reg_number = str(reg_number)
    

    if(choice==2):
        choice1 = int(input('Enter Choice:\n1 Login\n2. Register'))
        if(choice1==1):
            name=input("Login ID: ")
            dr_email_ID=input("Email ID: ")
            user=doctor_Login(name,dr_email_ID)
            user.dr_chk()
            print("logged in succesfully")
            name = name
            
        elif(choice1==2):
            name =input("Enter the name of the Doctor")
            email_id=input("Input email ID")
            phone_no=input("Input the contact number")
            address = input("Input the address")
            schedule=int(input("input the schedule you wish to choose, if your shift begins at 8AM write 800 and likewise"))
            age=input("input the age")
            password=input("input the password")
            position=input("input the position")
            specialization=input("input the specialization")
            department=input("input the department you wish to apply to")
            surgeon = input("Surgeon OR Senior Surgeon or Non Surgeon")
            timings2 = int(input("enter the time till which you are available, if your shift ends at 17:00 write 1700 and likewise "))
            #new=Doctor_registration(name,email_id,phone_no,address,schedule,age,password,position,specialization,department,surgeon)
            #new.save_doctor
            
            print(name)
            wb2 = load_workbook('Doctors_1.xlsx')
            cell1=wb2.active
            df = pd.DataFrame(cell1.values,columns=['name','email_id','phone_no','address','schedule','age','password','position','specialization','department','surgeon','timings2'])
            data1={'name':[name],'Email_id':[email_id],'phone_no':[phone_no],'address':[address],'schedule':[schedule],'age':[age],'password':[password],'position':[position],'specialization':[specialization],'department':[department],'surgeon':[surgeon],'timings2':[timings2]}
            df1=pd.DataFrame(data1,columns=['name','email_id','phone_no','address','schedule','age','password','position','specialization','department','surgeon','timings2'])
            df2=pd.concat([df,df1])
            writer = pd.ExcelWriter('Doctors_1.xlsx')
            df2.to_excel(writer,'Sheet1',index=False,index_label=False,header=False)
            writer.save()
            
        
        choice2 = int(input('Enter Choice:\n1 Details of a Patient\n2. Details of all the current patients'))
        if(choice2==1):
            choice21 = int(input('Enter Choice:\n1 List of Patients through registration number\n2. Search through patient name'))
            if(choice21==1):
                q = input("Enter the Registration number of the patient you wish to see the details of: ")
                df = Var2
                print(df.loc[df['Registration Number'] == str(q)])  
            else:
                q = input("Enter the name of patient whose details you wish to see: ")
                df = Var2
                df.loc[df['Patient Name'] == str(q)]
            choice22 = int(input("would you like to edit the details of patient ?: \n1 Yes\n2 No"))
            if(choice22==1):
                choice12344 = int(input("change details of \n1 existing patient \n2 New registration"))
                if (choice12344 == 1):
                    q = input("Enter the Registration number of the patient you wish to edit the details of: ")
                    df = Var2
                    df.loc[df['Registration Number'] == str(q)]
                if (choice12344 == 2):
                    reg_number = reg_number
                p_symptoms =input("Enter the symptoms of the patient")
                p_location = input("Input location to which the patient is being admitted: OPD or Emergency or Ward")
                Doctor_assigned= name
                prescription=input("Enter the medicines")
                Test=input("input the Test")
                Diagnosis=input("input the diagnosis")
                p_department=input("Enter the name of your department")
                p_entry_time= input("ENter the patient entry time")
                p_exit_time= input("ENter the patient exit time")
                
                choice222 = input(print("would you like to refer the patient to another doctor ? "))
                
                if (choice222== 'Yes' or 'YES' or 'yes' or 'yES'):
                    
                    df11 = Var1                    
                    #df11.loc[df11['Dr Name'] == str(d_chosen_department)]
                    #df[df['Dr']==5].index.values.astype(int)
                    index_self = df11[df11['Dr Name']==name].index.item()
                    self_post = Var1.iloc[index_self]['Post']
                    
                    self_department = Var1.iloc[index_self]['Department']
                    
                    if (self_post == 'JR' or 'SR'):
                        df11.loc[df11['Department'] == self_department]
                        #refered_doctor = 
                        #index_self1 = (df11.loc[(df11['Department']==self_department)] and df11.loc[df11['Post']=='specialist']).index.item()
                        print(df11.loc[(df11['Department']==self_department)])
                        chosen_doctor = int(input("enter the serial ID of the doctor you wish to refer to"))
                        if (Var1.iloc[chosen_doctor]['Post']== 'JR' or 'SR'):
                            print("sorry you can not refer this doctor")
                        if (Var1.iloc[chosen_doctor]['Post']== 'specialist' or 'super superspecialist'):
                            chosen_doctor = chosen_doctor
                    if (self_post == 'specialist'):
                        refered_doctor = int(input("enter the serial ID of the doctor you wish to refer to"))
                        print(df11.loc[(df11['Department']==self_department)])
                        if (Var1.iloc[refered_doctor]['Post']== 'JR' or 'SR' or 'specialist'):
                            print("sorry you can not refer this doctor")
                        if (Var1.iloc[refered_doctor]['Post']== 'super superspecialist'):
                            chosen_doctor = refered_doctor
                    if (self_post == 'super specialist'):
                        
                        choice412 = input('Enter choice:\n1 Neurology\n2 Cardiology\n3 ENT\n4 Orthopaedics\n5 Physician')
                        chosen_department1 = choice412
                        df11 = Var1
                        print(df11.loc[df11['Department'] == str(chosen_department)])
                    
                        qwe1 = input("Enter the serial number of the doctor you wish to choose")
                        chosen_doctor = Var1.iloc[int(qwe1)][0]
                        print("You have chosen Dr", chosen_doctor, "of", chosen_department1)
                chosen_department1 = p_department            
                chosen_doctor = Doctor_assigned
                wb2 = load_workbook('Patients_1.xlsx')
                cell1=wb2.active
                wb3 = load_workbook('Patients_1.xlsx')
                cell1=wb3.active
                df = pd.DataFrame(cell1.values,columns=['Patient Name','Registration Number','Age','email-ID','Contact Number','symptoms','location','Assigned Doctor','entry time','exit time','prescription','test','diagnosis','department'])
                data1={'Patient Name':[p_name],'Registration number':[reg_number],'Age':[p_age],'email-ID':[p_emailID],'Contact_Number':[p_contact_number],'symptoms':[p_symptoms],'location':[p_location],'Assigned Doctor':[chosen_doctor],'entry time':[p_entry_time],'exit time':[p_exit_time],'prescription':[prescription],'test':[Test],'diagnosis':[Diagnosis],'department':[chosen_department1]}
                df1=pd.DataFrame(data1,columns=['Patient Name','Registration Number','Age','email-ID','Contact Number','symptoms','location','Assigned Doctor','entry time','exit time','prescription','test','diagnosis','department'])
                df2=pd.concat([df,df1])
                writer = pd.ExcelWriter('Patients_1.xlsx')
                df2.to_excel(writer,'Sheet1',index=False,index_label=False,header=False)
                writer.save()
                print("You have been registered.")
                    
                    
                                   
                
        if(choice2==2):
            patients_assigned = str(name)
            df12 = Var2
            print(df12.loc[df12['Doctor_assigned'] == str(patients_assigned)])
            
                
           
           
           
'''           
    Var1.iloc[2][3]
        
    i = MRD1001
    Var2[:2]     
    Var2.ix['MRD1001']
    df = Var2
    df.loc[df['Registration Number'] == 'MRD1001']        
    
            
          Var1  
          Var2

          user_name=input("Enter Name: ")
            Age=input("Age: ")
            contact=input("contact: ")
            new=user_registration(user_name,Age,contact)
            new.save_user()
'''
'''
            
            
'''          
            
            
    #choice=int(input("Enter Choice:\n1.Display doctor details\n2.Patient Registration\n3.Delete Patient\n4.Select Doctor\n5.Pay Bill\n6.Display Admin Details\n7.Display Support Staff Details\n8.Doctor's Account\n9.Admin's Account"))
    
   
main()        
